# -*- coding: utf-8 -*-
"""按月度汇总模版生成发票汇总表。

用法:
    python gen_summary.py --data data.json --source-dir 源发票目录 --output 月度汇总.xlsx

data.json 结构:
{
  "month": "2026年7月",
  "records": [
    {
      "name": "黄卉",
      "category": "差旅费",
      "number": "26378324211049524479",
      "date": "7月6日",
      "amount": "940.00",
      "description": "机票 北京-厦门 SC2130"
    }
  ]
}
"""
import argparse
import copy
import json
import os
import re
import sys
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_TEMPLATE = os.path.join(SCRIPT_DIR, '..', 'assets', '月度汇总模版.xlsx')
CATEGORIES = ('餐饮费', '办公费', '差旅费', '交通费', '快递费')


def amount(value, field):
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        sys.exit('%s 不是有效金额: %r' % (field, value))


def date_key(value):
    text = str(value)
    match = re.search(r'(\d+)月(\d+)日', text)
    if match:
        return int(match.group(1)), int(match.group(2))
    match = re.search(r'(?:\d{4}[-/]?)?(\d{1,2})[-/](\d{1,2})', text)
    return (int(match.group(1)), int(match.group(2))) if match else (float('inf'), float('inf'))


def validate(records):
    required = ('name', 'category', 'number', 'date', 'amount', 'description')
    for i, record in enumerate(records, 1):
        missing = [key for key in required if key not in record]
        if missing:
            sys.exit('第 %d 条记录缺少字段: %s' % (i, ', '.join(missing)))
        if record['category'] not in CATEGORIES:
            sys.exit('第 %d 条记录类别无效: %s' % (i, record['category']))
        amount(record['amount'], '第 %d 条记录金额' % i)


def write_row(ws, row, values, styles):
    for column, value in enumerate(values, 1):
        cell = ws.cell(row, column)
        if cell.__class__.__name__ != 'MergedCell':
            cell._style = copy.copy(styles[column - 1])
            cell.value = value


def main():
    parser = argparse.ArgumentParser(description='按模版生成发票月度汇总表')
    parser.add_argument('--data', required=True, help='数据 JSON 文件路径')
    parser.add_argument('--source-dir', required=True, help='源发票目录; 生成文件直接写入该目录')
    parser.add_argument('--output', required=True, help='输出文件名, 只能是当前目录下的 .xlsx 文件名')
    parser.add_argument('--template', default=DEFAULT_TEMPLATE, help='模版 .xlsx 路径')
    args = parser.parse_args()
    if Path(args.output).name != args.output or Path(args.output).suffix.lower() != '.xlsx':
        sys.exit('--output 必须是当前目录下的 .xlsx 文件名, 不得包含子目录')
    output = os.path.join(args.source_dir, args.output)

    with open(args.data, encoding='utf-8') as f:
        data = json.load(f)
    if not isinstance(data.get('month'), str) or not data['month']:
        sys.exit('data.json 的 month 必须是非空字符串')
    records = data.get('records')
    if not isinstance(records, list) or not records:
        sys.exit('data.json 的 records 必须是非空列表')
    validate(records)

    groups = defaultdict(list)
    for record in records:
        groups[(record['name'], record['category'])].append(record)
    names = sorted({record['name'] for record in records})
    ordered_groups = []
    for name in names:
        for category in CATEGORIES:
            group = groups.get((name, category))
            if group:
                group.sort(key=lambda r: (date_key(r['date']), amount(r['amount'], '金额'), str(r['number'])))
                total = sum((amount(r['amount'], '金额') for r in group), Decimal(0))
                ordered_groups.append((name, category, group, total))

    wb = load_workbook(args.template)
    ws = wb.active
    title_styles = [copy.copy(ws.cell(1, column)._style) for column in range(1, 5)]
    person_styles = [copy.copy(ws.cell(3, column)._style) for column in range(1, 5)]
    category_styles = [copy.copy(ws.cell(4, column)._style) for column in range(1, 5)]
    header_styles = [copy.copy(ws.cell(5, column)._style) for column in range(1, 5)]
    data_styles = [copy.copy(ws.cell(6, column)._style) for column in range(1, 5)]
    for merged_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_range))
    ws.delete_rows(1, ws.max_row)
    ws.merge_cells('A1:D1')
    write_row(ws, 1, [data['month'] + ' 发票月度汇总表', None, None, None], title_styles)
    ws.row_dimensions[1].height = 28

    row = 3
    previous_name = None
    for name, category, group, total in ordered_groups:
        if previous_name is not None and name != previous_name:
            row += 1
        if name != previous_name:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            write_row(ws, row, [name, None, None, None], person_styles)
            ws.row_dimensions[row].height = 24
            row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        write_row(ws, row, ['%s（%.2f元）' % (category, total), None, None, None], category_styles)
        row += 1
        write_row(ws, row, ['发票编号', '日期', '总金额(元)', '简要说明'], header_styles)
        row += 1
        for record in group:
            write_row(ws, row, [record['number'], record['date'], '%.2f' % amount(record['amount'], '金额'), record['description']], data_styles)
            row += 1
        previous_name = name

    os.makedirs(os.path.abspath(args.source_dir), exist_ok=True)
    wb.save(output)
    print('已生成:', output)


if __name__ == '__main__':
    main()
